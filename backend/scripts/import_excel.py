"""
import_excel.py — reads Spark_Schema_1.xlsx seed data and inserts into Supabase.

Usage:
  cd Claude_Code
  python backend/scripts/import_excel.py --file .venv/RefDoc/Spark_Schema_1.xlsx

Expects sheets: subjects, topics, topic_content, real_world_apps,
                experts, careers, topic_careers, career_skills, expert_videos,
                quiz_questions, quiz_options
"""

import argparse
import sys
import openpyxl
from supabase import create_client
import os
from dotenv import load_dotenv

load_dotenv("backend/.env")

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_KEY = os.environ["SUPABASE_SERVICE_KEY"]
sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


def sheet_to_dicts(ws):
    """Convert a worksheet to a list of dicts using the first row as headers."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else None for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {headers[i]: v for i, v in enumerate(row) if headers[i]}
        result.append(d)
    return result


def upsert(table, rows, conflict_col="id"):
    if not rows:
        print(f"  ⚠ No rows for {table}")
        return
    res = sb.table(table).upsert(rows).execute()
    print(f"  ✓ {table}: {len(res.data)} rows upserted")


def run(filepath):
    print(f"Loading {filepath} …")
    wb = openpyxl.load_workbook(filepath)
    available = wb.sheetnames
    print(f"Sheets found: {available}\n")

    # ── subjects ──
    if "subjects" in available:
        rows = sheet_to_dicts(wb["subjects"])
        upsert("subjects", rows, "subject_id")

    # ── topics ──
    if "topics" in available:
        rows = sheet_to_dicts(wb["topics"])
        upsert("topics", rows, "topic_id")

    # ── topic_content ──
    if "topic_content" in available:
        rows = sheet_to_dicts(wb["topic_content"])
        upsert("topic_content", rows, "id")

    # ── real_world_apps ──
    if "real_world_apps" in available:
        rows = sheet_to_dicts(wb["real_world_apps"])
        upsert("real_world_apps", rows, "id")

    # ── experts ──
    if "experts" in available:
        rows = sheet_to_dicts(wb["experts"])
        upsert("experts", rows, "expert_id")

    # ── careers ──
    if "careers" in available:
        rows = sheet_to_dicts(wb["careers"])
        upsert("careers", rows, "career_id")

    # ── topic_careers (junction) ──
    if "topic_careers" in available:
        rows = sheet_to_dicts(wb["topic_careers"])
        # composite PK: upsert handles it
        sb.table("topic_careers").upsert(rows).execute()
        print(f"  ✓ topic_careers: {len(rows)} rows upserted")

    # ── career_skills ──
    if "career_skills" in available:
        rows = sheet_to_dicts(wb["career_skills"])
        upsert("career_skills", rows, "id")

    # ── expert_videos ──
    if "expert_videos" in available:
        rows = sheet_to_dicts(wb["expert_videos"])
        upsert("expert_videos", rows, "id")

    # ── quiz_questions ──
    if "quiz_questions" in available:
        rows = sheet_to_dicts(wb["quiz_questions"])
        upsert("quiz_questions", rows, "question_id")

    # ── quiz_options ──
    if "quiz_options" in available:
        rows = sheet_to_dicts(wb["quiz_options"])
        upsert("quiz_options", rows, "option_id")

    print("\nDone.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Path to Spark_Schema_1.xlsx")
    args = parser.parse_args()
    run(args.file)
